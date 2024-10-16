
import pandas as pd
import os
from datetime import datetime
import openpyxl
filepath = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'pool.xlsx')

def get_projects():
    projects = pd.read_excel(filepath, sheet_name='Projects')
    projects['vid'] = projects.index # vid is virtual id. its used as a unique identifier for the tables
    projects['pickup_date'] = pd.to_datetime(projects['pickup_date'])
    projects['return_date'] = pd.to_datetime(projects['return_date'])
    return projects.copy()

def get_projects_table():
    projects = get_projects()
    projects['pickup_date'] = projects['pickup_date'].dt.strftime('%Y-%m-%d').astype(str)
    projects['return_date'] = projects['return_date'].dt.strftime('%Y-%m-%d').astype(str)
    return projects

def get_projects_timeline():
    projects_timeline = get_projects_table()
    projects_timeline['repeat'] = projects_timeline.groupby('Projects').cumcount() + 1
    projects_timeline['Name'] = projects_timeline.apply(lambda x: f"{x['Projects']} {x['repeat']-1}" if x['repeat'] > 1 else x['Projects'], axis=1)
    projects_timeline['Name'] = projects_timeline['Name'].apply(lambda x: x[:12] + '...' if len(x) > 0 else x)
    return projects_timeline

def get_inventory():
    inventory = pd.read_excel(filepath, sheet_name='Inventory')
    inventory['vid'] = inventory.index # vid is virtual id. its used as a unique identifier for the tables
    return inventory.copy()

def get_inventory_instruments_number():
    inventory = get_inventory()
    summarized_inventory = inventory.groupby('ID', as_index=False).agg({
        'Number_sum': 'sum',
        'Instrument name': 'first',  # or you could use 'last' or 'max' depending on your requirement
        'Type': 'first'
    })
    return summarized_inventory

def get_repairs():
    repairs = pd.read_excel(filepath, sheet_name='Inventory_Repair')
    repairs['vid'] = repairs.index # vid is virtual id. its used as a unique identifier for the tables
    return repairs.copy()

def get_datepicker_dates():
    projects = get_projects()
    earliest_pickup_date = projects['pickup_date'].min()
    latest_return_date = projects['return_date'].max()
    earliest_pickup_date = earliest_pickup_date - pd.DateOffset(months=1)
    latest_return_date = latest_return_date + pd.DateOffset(years=1)
    today = datetime.now()
    return earliest_pickup_date.strftime('%Y-%m-%d'), latest_return_date.strftime('%Y-%m-%d'), today.strftime('%Y-%m-%d') 

def add_to_table(data, sheet):
    try:
        # Check if the file exists before trying to read it
        if os.path.exists(filepath):
            # Load the existing 'sheet' sheet into a DataFrame
            dataframe = pd.read_excel(filepath, sheet_name=sheet, engine='openpyxl')
        else:
            # If the file doesn't exist, create an empty DataFrame with the columns similar to project_data
            dataframe = pd.DataFrame(columns=data.keys())
        
        # Convert project_data (dictionary) into a DataFrame row
        new_row = pd.DataFrame([data])
        
        # Append the new project row to the existing DataFrame
        updated_dataframe = pd.concat([dataframe, new_row], ignore_index=True)
        # Save the updated DataFrame back to the same Excel file
        with pd.ExcelWriter(filepath, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            updated_dataframe.to_excel(writer, sheet_name=sheet, index=False)
        # Success response
        return {"status": "success", "message": f"New row added to {sheet} successfully!"}
    
    except Exception as e:
        # Error response in case of any failure
        return {"status": "error", "message": f"An error occurred: {str(e)}"}
    

def update_table(changes, sheetname):
    """
    Apply changes (update or delete) to a specific Excel sheet based on rowIndex.

    Args:
        changes (list): List of dictionaries, where each dictionary contains:
                        {"rowIndex": row_id, "status": "update" or "delete", "row": updated_row}.
        sheetname (str): Name of the sheet in the Excel file to update.
    
    Returns:
        dict: A dictionary with "status" and "message" to indicate success or failure.
    """
    try:
        # Load the workbook and get the specific sheet
        workbook = openpyxl.load_workbook(filepath)
        if sheetname not in workbook.sheetnames:
            return {"status": "error", "message": f"Sheet '{sheetname}' not found in the Excel file."}
        sheet = workbook[sheetname]

        # Apply each change from the "changes" list
        for change in changes:
            row_index = change["rowIndex"] + 2  # Excel rows are 1-based, and usually row 1 is headers
            status = change["status"]
            updated_row = change["row"]

            if status == "update":
                # Update the existing row with new values
                for col, (key, value) in enumerate(updated_row.items(), start=1):
                    # Check if the column is 'pickup_date' or 'return_date' to maintain the format
                    if key in ['pickup_date', 'return_date']:
                        original_cell = sheet.cell(row=row_index, column=col)
                        # if isinstance(original_cell.value, datetime):
                        #     # If the original cell is a datetime object, convert the new value to datetime
                        #     value = datetime.strptime(value, '%Y-%m-%d') if isinstance(value, str) else value
                        if isinstance(original_cell.value, datetime):
                            # If the original cell is a datetime object, convert the new value to datetime
                            if isinstance(value, str):
                                # Parse the ISO 8601 format
                                value = datetime.fromisoformat(value.replace("T", " "))
                            
                            # Update the value to maintain the same shape as original_cell
                            value = value.replace(
                                hour=original_cell.value.hour,
                                minute=original_cell.value.minute,
                                second=original_cell.value.second,
                                microsecond=original_cell.value.microsecond,
                                tzinfo=original_cell.value.tzinfo  # Keep the original timezone info
                            )
                        # Apply the value with the same formatting
                        sheet.cell(row=row_index, column=col).value = value
                        sheet.cell(row=row_index, column=col).number_format = original_cell.number_format  # Maintain original format
                    else:
                        # Update non-date fields as normal
                        sheet.cell(row=row_index, column=col).value = value
                print(f"Updated row {row_index} with data: {updated_row}")

            elif status == "delete":
                # Delete the entire row
                sheet.delete_rows(row_index)
                print(f"Deleted row {row_index}")

        # Save the updated workbook
        workbook.save(filepath)
        return {"status": "success", "message": "Excel file updated successfully."}

    except FileNotFoundError:
        return {"status": "error", "message": f"File not found: {filepath}"}
    
    except PermissionError:
        return {"status": "error", "message": f"Permission denied: {filepath}. Close the file if it's open."}
    
    except Exception as e:
        return {"status": "error", "message": f"An error occurred: {str(e)}"}
